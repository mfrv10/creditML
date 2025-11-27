"""
Report Generator
Creates comprehensive Excel and PDF reports from analysis results
"""

import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from typing import Dict

class ReportGenerator:
    """
    Generate professional reports from credit analysis
    """
    
    def __init__(self, df: pd.DataFrame, metrics: Dict, valuation: Dict):
        """
        Initialize report generator
        
        Args:
            df: Analyzed portfolio DataFrame
            metrics: Dictionary of calculated metrics
            valuation: Dictionary of valuation metrics
        """
        self.df = df
        self.metrics = metrics
        self.valuation = valuation
        self.timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    def create_excel_report(self) -> BytesIO:
        """
        Create comprehensive Excel report with multiple sheets including cash flow analysis
        
        Returns:
            BytesIO buffer containing Excel file
        """
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Sheet 1: Executive Summary
            self._create_summary_sheet(writer)
            
            # Sheet 2: Scored Portfolio
            self._create_portfolio_sheet(writer)
            
            # Sheet 3: Risk Analysis
            self._create_risk_sheet(writer)
            
            # Sheet 4: Financial Ratios
            self._create_ratios_sheet(writer)
            
            # Sheet 5: High Risk Accounts
            self._create_highrisk_sheet(writer)
            
            # NEW: Sheet 6 & 7: Cash Flow Analysis (both methods)
            if 'cash_flow_analysis' in self.valuation:
                self._create_cashflow_comparison_sheet(writer)
                self._create_cashflow_details_sheet(writer)
        
        buffer.seek(0)
        return buffer
    
    def _create_summary_sheet(self, writer):
        """Create executive summary sheet"""
        summary_data = {
            'Metric': [
                'Report Generated',
                'Total Accounts',
                'Total Exposure',
                'Total Outstanding Balance',
                'Average Credit Limit',
                'Average Utilization',
                '',
                'Portfolio Risk Metrics',
                'Average Default Probability',
                'High Risk Accounts',
                'Expected Loss',
                'Value at Risk (95%)',
                '',
                'Portfolio Valuation',
                'Book Value',
                'Risk-Adjusted Value',
                'Expected Revenue',
                'Net Portfolio Value',
                'Return on Assets'
            ],
            'Value': [
                self.timestamp,
                f"{len(self.df):,}",
                f"${self.metrics['total_exposure']:,.0f}",
                f"${self.metrics['total_outstanding']:,.0f}",
                f"${self.metrics['avg_credit_limit']:,.0f}",
                f"{self.metrics['avg_utilization']:.2%}",
                '',
                '',
                f"{self.metrics['avg_default_probability']:.2%}",
                f"{self.metrics['high_risk_count']:,}",
                f"${self.metrics['expected_loss']:,.0f}",
                f"${self.metrics['var_95']:,.0f}",
                '',
                '',
                f"${self.valuation['book_value']:,.0f}",
                f"${self.valuation['risk_adjusted_value']:,.0f}",
                f"${self.valuation['expected_revenue']:,.0f}",
                f"${self.valuation['net_value']:,.0f}",
                f"{self.valuation['roa']:.2%}"
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
        
        # Format the sheet
        ws = writer.sheets['Executive Summary']
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
    
    def _create_portfolio_sheet(self, writer):
        """Create detailed portfolio data sheet"""
        # Select relevant columns
        portfolio_cols = [
            'account_id', 'credit_limit', 'outstanding_balance', 
            'utilization_rate', 'default_probability', 'risk_score', 
            'risk_category'
        ]
        
        available_cols = [c for c in portfolio_cols if c in self.df.columns]
        portfolio_df = self.df[available_cols].copy()
        
        # Format numbers
        if 'default_probability' in portfolio_df.columns:
            portfolio_df['default_probability'] = portfolio_df['default_probability'].round(4)
        if 'utilization_rate' in portfolio_df.columns:
            portfolio_df['utilization_rate'] = portfolio_df['utilization_rate'].round(4)
        if 'risk_score' in portfolio_df.columns:
            portfolio_df['risk_score'] = portfolio_df['risk_score'].round(2)
        
        portfolio_df.to_excel(writer, sheet_name='Portfolio Details', index=False)
        
        # Format sheet
        ws = writer.sheets['Portfolio Details']
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
    
    def _create_risk_sheet(self, writer):
        """Create risk analysis sheet"""
        # Risk distribution by category
        risk_dist = self.df['risk_category'].value_counts().reset_index()
        risk_dist.columns = ['Risk Category', 'Count']
        risk_dist['Percentage'] = (risk_dist['Count'] / len(self.df) * 100).round(2)
        
        # Risk by utilization bands
        self.df['utilization_band'] = pd.cut(
            self.df['utilization_rate'],
            bins=[0, 0.3, 0.5, 0.7, 0.9, 2.0],
            labels=['0-30%', '30-50%', '50-70%', '70-90%', '>90%']
        )
        risk_by_util = self.df.groupby('utilization_band').agg({
            'account_id': 'count',
            'default_probability': 'mean'
        }).reset_index()
        risk_by_util.columns = ['Utilization Band', 'Count', 'Avg Default Prob']
        
        # Write to Excel with spacing
        risk_dist.to_excel(writer, sheet_name='Risk Analysis', index=False, startrow=0)
        
        # Add risk by utilization below
        start_row = len(risk_dist) + 3
        ws = writer.sheets['Risk Analysis']
        ws.cell(start_row, 1, 'Risk by Utilization Band').font = Font(bold=True, size=12)
        risk_by_util.to_excel(writer, sheet_name='Risk Analysis', index=False, startrow=start_row+1)
    
    def _create_ratios_sheet(self, writer):
        """Create financial ratios sheet"""
        ratios_data = []
        
        for ratio_name, ratio_value in self.metrics['ratios'].items():
            if isinstance(ratio_value, float):
                formatted_value = f"{ratio_value:.4f}" if abs(ratio_value) < 10 else f"{ratio_value:,.2f}"
            else:
                formatted_value = str(ratio_value)
            
            ratios_data.append({
                'Ratio': ratio_name,
                'Value': formatted_value
            })
        
        ratios_df = pd.DataFrame(ratios_data)
        ratios_df.to_excel(writer, sheet_name='Financial Ratios', index=False)
        
        # Format sheet
        ws = writer.sheets['Financial Ratios']
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def _create_highrisk_sheet(self, writer):
        """Create high-risk accounts sheet"""
        high_risk = self.df.nlargest(50, 'default_probability').copy()
        
        # Select relevant columns
        cols = ['account_id', 'credit_limit', 'outstanding_balance', 
                'utilization_rate', 'default_probability', 'risk_score']
        available_cols = [c for c in cols if c in high_risk.columns]
        
        high_risk[available_cols].to_excel(
            writer, 
            sheet_name='High Risk Accounts', 
            index=False
        )
        
        # Format sheet
        ws = writer.sheets['High Risk Accounts']
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
    
    def _create_cashflow_comparison_sheet(self, writer):
        """Create cash flow method comparison sheet"""
        cf_analysis = self.valuation['cash_flow_analysis']
        
        # Get metrics from both methods
        method_a = cf_analysis['method_a']
        method_b = cf_analysis['method_b']
        comparison = cf_analysis['comparison']
        
        # Create comparison summary
        comparison_data = {
            'Metric': [
                'VALUATION METHOD',
                '',
                'NPV (Net Present Value)',
                'IRR (Internal Rate of Return)',
                'MOIC (Multiple on Invested Capital)',
                'Payback Period (months)',
                '',
                'CASH FLOWS',
                'Total Collections (24 months)',
                'Total Losses (24 months)',
                'Net Cash Flow',
                '',
                'RETURN METRICS',
                'ROI %',
                'Discount Rate Used',
                'Break-even Loss Rate'
            ],
            'Method A: Static Pool Analysis': [
                'Revolving Credit Approach',
                '',
                f"${method_a['npv']:,.0f}",
                f"{method_a['irr_annual']:.2f}%",
                f"{method_a['moic']:.2f}x",
                f"{method_a['payback_period_months']:.1f}",
                '',
                '',
                f"${method_a['total_collections']:,.0f}",
                f"${method_a['total_losses']:,.0f}",
                f"${method_a['total_collections'] - method_a['total_losses']:,.0f}",
                '',
                '',
                f"{method_a['roi_percent']:.2f}%",
                f"{method_a['discount_rate_used']:.0f}%",
                f"{method_a['breakeven_loss_rate']*100:.2f}%"
            ],
            'Method B: Dynamic CF Model': [
                'Term Loan Approach',
                '',
                f"${method_b['npv']:,.0f}",
                f"{method_b['irr_annual']:.2f}%",
                f"{method_b['moic']:.2f}x",
                f"{method_b['payback_period_months']:.1f}",
                '',
                '',
                f"${method_b['total_collections']:,.0f}",
                f"${method_b['total_losses']:,.0f}",
                f"${method_b['total_collections'] - method_b['total_losses']:,.0f}",
                '',
                '',
                f"{method_b['roi_percent']:.2f}%",
                f"{method_b['discount_rate_used']:.0f}%",
                f"{method_b['breakeven_loss_rate']*100:.2f}%"
            ],
            'Difference': [
                '',
                '',
                f"${method_a['npv'] - method_b['npv']:,.0f}",
                f"{method_a['irr_annual'] - method_b['irr_annual']:.2f}pp",
                f"{method_a['moic'] - method_b['moic']:.2f}x",
                f"{method_a['payback_period_months'] - method_b['payback_period_months']:.1f}",
                '',
                '',
                f"${comparison['total_collections']['Difference']:,.0f}",
                f"${comparison['total_losses']['Difference']:,.0f}",
                '',
                '',
                '',
                f"{method_a['roi_percent'] - method_b['roi_percent']:.2f}pp",
                '',
                ''
            ]
        }
        
        comparison_df = pd.DataFrame(comparison_data)
        comparison_df.to_excel(writer, sheet_name='CF Valuation Comparison', index=False)
        
        # Format sheet
        ws = writer.sheets['CF Valuation Comparison']
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        
        # Bold and color header row
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Bold section headers
        for row in [2, 9, 14]:
            ws.cell(row, 1).font = Font(bold=True, size=11)
        
        # Add methodology explanation below
        start_row = len(comparison_df) + 3
        ws.cell(start_row, 1, 'KEY DIFFERENCES:').font = Font(bold=True, size=12)
        
        explanations = [
            ['Method A (Static Pool)', 'Best for revolving credit (credit cards), regulatory reporting, quick analysis'],
            ['Method B (Dynamic CF)', 'Best for term loans (auto, mortgage), M&A valuation, detailed analysis'],
            ['', ''],
            ['Method A Approach', 'Aggregates portfolio into risk cohorts, models cohort behavior over time'],
            ['Method B Approach', 'Models each loan individually with specific amortization schedules'],
            ['', ''],
            ['Computational Cost', 'Method A: Low (fast), Method B: Higher (more detailed)'],
            ['Accuracy Trade-off', 'Method A: Good for portfolios, Method B: Precise for individual loans']
        ]
        
        for i, exp in enumerate(explanations):
            ws.cell(start_row + i + 1, 1, exp[0]).font = Font(bold=True)
            ws.cell(start_row + i + 1, 2, exp[1])
    
    def _create_cashflow_details_sheet(self, writer):
        """Create detailed monthly cash flow projections"""
        cf_analysis = self.valuation['cash_flow_analysis']
        
        # Get monthly comparison
        monthly_comp = cf_analysis['comparison']['monthly_comparison']
        
        # Format for Excel
        monthly_comp['Month'] = monthly_comp['Month'].astype(int)
        
        monthly_comp.to_excel(writer, sheet_name='Monthly Cash Flows', index=False)
        
        # Format sheet
        ws = writer.sheets['Monthly Cash Flows']
        
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Add summary below the table
        start_row = len(monthly_comp) + 3
        
        ws.cell(start_row, 1, 'SUMMARY STATISTICS').font = Font(bold=True, size=12)
        
        summary_stats = [
            ['', 'Static Pool', 'Dynamic CF'],
            ['Total Collections', f"${monthly_comp['Static Pool Collections'].sum():,.0f}", 
             f"${monthly_comp['Dynamic CF Collections'].sum():,.0f}"],
            ['Total Losses', f"${monthly_comp['Static Pool Losses'].sum():,.0f}", 
             f"${monthly_comp['Dynamic CF Losses'].sum():,.0f}"],
            ['Average Monthly CF', f"${monthly_comp['Static Pool Net CF'].mean():,.0f}", 
             f"${monthly_comp['Dynamic CF Net CF'].mean():,.0f}"],
            ['Peak Monthly CF', f"${monthly_comp['Static Pool Net CF'].max():,.0f}", 
             f"${monthly_comp['Dynamic CF Net CF'].max():,.0f}"]
        ]
        
        for i, row_data in enumerate(summary_stats):
            for j, cell_value in enumerate(row_data):
                cell = ws.cell(start_row + i + 1, j + 1, cell_value)
                if i == 0:
                    cell.font = Font(bold=True)
    
    def create_pdf_report(self) -> BytesIO:
        """
        Create executive PDF summary report
        
        Returns:
            BytesIO buffer containing PDF file
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12
        )
        
        # Title
        story.append(Paragraph("Credit Portfolio Analysis Report", title_style))
        story.append(Paragraph(f"Generated: {self.timestamp}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Accounts', f"{len(self.df):,}"],
            ['Total Exposure', f"${self.metrics['total_exposure']:,.0f}"],
            ['Total Outstanding', f"${self.metrics['total_outstanding']:,.0f}"],
            ['Average Utilization', f"{self.metrics['avg_utilization']:.2%}"],
            ['Expected Loss', f"${self.metrics['expected_loss']:,.0f}"],
            ['High Risk Accounts', f"{self.metrics['high_risk_count']:,}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Risk Distribution
        story.append(Paragraph("Risk Distribution", heading_style))
        
        risk_dist = self.df['risk_category'].value_counts()
        risk_data = [['Risk Category', 'Count', 'Percentage']]
        for cat, count in risk_dist.items():
            risk_data.append([cat, f"{count:,}", f"{count/len(self.df)*100:.1f}%"])
        
        risk_table = Table(risk_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        risk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(risk_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Portfolio Valuation
        story.append(Paragraph("Portfolio Valuation", heading_style))
        
        val_data = [
            ['Metric', 'Value'],
            ['Book Value', f"${self.valuation['book_value']:,.0f}"],
            ['Expected Loss', f"${self.valuation['expected_loss']:,.0f}"],
            ['Risk-Adjusted Value', f"${self.valuation['risk_adjusted_value']:,.0f}"],
            ['Expected Revenue', f"${self.valuation['expected_revenue']:,.0f}"],
            ['Net Portfolio Value', f"${self.valuation['net_value']:,.0f}"],
            ['Return on Assets', f"{self.valuation['roa']:.2%}"],
        ]
        
        val_table = Table(val_data, colWidths=[3*inch, 2*inch])
        val_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(val_table)
        
        # Add Cash Flow Valuation if available
        if 'cash_flow_analysis' in self.valuation:
            story.append(Spacer(1, 0.3*inch))
            story.append(Paragraph("Cash Flow-Based Valuation", heading_style))
            
            cf_analysis = self.valuation['cash_flow_analysis']
            method_a = cf_analysis['method_a']
            method_b = cf_analysis['method_b']
            
            cf_val_data = [
                ['Metric', 'Method A (Static Pool)', 'Method B (Dynamic CF)'],
                ['NPV', f"${method_a['npv']:,.0f}", f"${method_b['npv']:,.0f}"],
                ['IRR', f"{method_a['irr_annual']:.2f}%", f"{method_b['irr_annual']:.2f}%"],
                ['MOIC', f"{method_a['moic']:.2f}x", f"{method_b['moic']:.2f}x"],
                ['Payback', f"{method_a['payback_period_months']:.0f} months", 
                 f"{method_b['payback_period_months']:.0f} months"],
            ]
            
            cf_val_table = Table(cf_val_data, colWidths=[2*inch, 1.75*inch, 1.75*inch])
            cf_val_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9)
            ]))
            
            story.append(cf_val_table)
            story.append(Spacer(1, 0.2*inch))
            
            # Add methodology note
            method_note = Paragraph(
                "<b>Method A:</b> Static Pool Analysis - Models aggregate cohort behavior (best for credit cards)<br/>"
                "<b>Method B:</b> Dynamic Cash Flow - Models individual loan amortization (best for term loans)",
                styles['Normal']
            )
            story.append(method_note)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer
